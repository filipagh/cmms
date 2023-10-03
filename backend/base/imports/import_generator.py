from datetime import datetime

from openpyxl import load_workbook


def parse_road_segment(text):
    parts = text.split()

    # Check if the text can be split into exactly 3 parts
    if len(parts) >= 3:
        prefix = parts[0]
        number = int(parts[1])  # convert to integer
        name = " ".join(parts[2:])
        return (prefix, number, name)
    else:
        return None


# Load the workbook
stations = load_workbook('Meteostanice _NDS_suradnice.xlsx')

sheet = stations["List1"]
rs_map = {}
station_map = {}
station_count = 0


# Get a specific cell
def parse_station(x, rs_id):
    global station_count
    station_name = sheet.cell(x, 1).value.strip()
    if station_name is None:
        return
    nord = sheet.cell(x, 2).value
    east = sheet.cell(x, 3).value
    try:
        description = ""
        km = int(sheet.cell(x, 5).value)
    except ValueError:
        km = 9999
        description = sheet.cell(x, 5).value

    station_id = "s" + str(station_count)
    station_count += 1
    station_map[station_name] = station_id

    print(f'''    {station_id} = station_service.create_station(
        StationNewSchema(name="{station_name}", road_segment_id={rs_id}, km_of_road={km},
                         km_of_road_note="",
                         latitude={str(nord)}, longitude={str(east)}, see_level=None, description="{description}"))''')


for x in range(5, sheet.max_row + 1):
    cell_value = sheet.cell(x, 4).value
    if cell_value is None:
        continue

    parsed_value = parse_road_segment(cell_value)
    if parsed_value is None:
        continue
    rs_prefix, rs_number, rs_name = parsed_value

    if rs_name not in rs_map:
        rs_map[rs_name] = "rs" + str(rs_number)
        print(
            f'    {rs_map[rs_name]} = rs_router.create_road_segment(RoadSegmentNewSchema(name="{rs_name}", ssud="{rs_prefix} {rs_number}"))')

    parse_station(x, rs_map[rs_name])


def parse_date(date):
    return "datetime(" + date.strftime("%Y,%-m,%-d") + ")"


def parse_components(sheet):
    for x in range(2, sheet.max_row):

        station_name = sheet.cell(x, 2).value.strip()
        station_id = station_map[station_name]
        component_name = sheet.cell(x, 3).value.strip()
        install_date = sheet.cell(x, 4).value
        serial_number = sheet.cell(x, 5).value
        component_waranty_date = sheet.cell(x, 6).value
        prepaid_service = sheet.cell(x, 7).value

        component_waranty_id = "None"
        if component_waranty_date is None:
            waranty = "NAN"
        elif prepaid_service is None:
            waranty = "COMPANY_WARRANTY"
        else:
            if prepaid_service < datetime.now():
                waranty = "COMPANY_WARRANTY"
            else:
                waranty = "INVESTMENT_CONTRACT"
                component_waranty_id = 'inv_id'

        if install_date is not None:
            install_date = parse_date(install_date)
        else:
            raise Exception("Install date is None on row " + str(x))

        if component_waranty_date is not None:
            component_waranty_date = parse_date(component_waranty_date)
        else:
            component_waranty_date = "None"

        if prepaid_service is not None:
            prepaid_service = parse_date(prepaid_service)
        else:
            prepaid_service = "None"

        print(f'''    assigned_component_rest_router.create_installed_component(
            new_components=_get_selected_assets_to_install(
                asset_ids=[_InstallAsset({component_name.replace(" ", "_").replace("-", "_").replace(",", "_").replace(".", "_").replace("(", "_").replace(")", "_")}, "{serial_number if serial_number is not None else ""}")],
                station_id={station_id}),
            component_warranty_id={component_waranty_id},
            components_warranty_source=ComponentWarrantySource.{waranty},
            component_warranty_until={component_waranty_date}, paid_service_until={prepaid_service},
            installation_date={install_date})
        ''')


stations = load_workbook('Filip_Stanice_komponenty.xlsx')

sheet = stations["Behárovce_SSÚD10"]
parse_components(sheet)

sheet = stations["Zvolen_SSÚR3"]
parse_components(sheet)

sheet = stations["Nová Baňa_SSÚR2"]
parse_components(sheet)

sheet = stations["Považská Bystrica_SSÚD5"]
parse_components(sheet)
